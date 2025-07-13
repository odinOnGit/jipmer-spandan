from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from django.utils import timezone
from django.http import HttpResponse
from .models import DelegateCardRegistration, EventRegistration, PassPurchase
from .serializers import DelegateCardRegistrationSerializer, EventRegistrationSerializer, PassPurchaseSerializer
import openpyxl
from openpyxl.utils import get_column_letter
from registration.utils.email import send_smtp_email
import json
from django.db import transaction
from rest_framework.parsers import MultiPartParser, FormParser

# Constants
EXEMPT_EVENTS = {
    'snap_saga', 'verse', 'reel_realm',
    'junior_mediquiz', 'senior_mediquiz',
    'case_presentation', 'poster_presentation',
    'research_paper', 'mediquiz', 'clinico_pathological_conference',
}

SPORTS_EVENTS = {'football', 'cricket', 'badminton', 'volleyball', 'table_tennis', 'throwball'}
CULT_EVENTS = {'solo_dance', 'duet_dance', 'mono_act', 'street_play', 'meme_making', 'face_painting'}
CULT_MAJOR_EVENTS = {'chorea', 'alaap', 'tinnitus', 'dernier_cri'}
LIT_EVENTS = {'debate', 'poetry', 'essay', 'elocution'}
QUIZ_EVENTS = {'medical_quiz', 'general_quiz'}

# Delegate Card Views
class DelegateCardRegisterView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = DelegateCardRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            reg = serializer.save()
            return Response({
                "message": "Registration submitted",
                "registration": DelegateCardRegistrationSerializer(reg).data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        self.check_permissions(request)
        regs = DelegateCardRegistration.objects.all()
        serializer = DelegateCardRegistrationSerializer(regs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def verify_registration(request, pk):
    try:
        reg = DelegateCardRegistration.objects.get(pk=pk)
    except DelegateCardRegistration.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
    
    reg.is_verified = True
    reg.status = 'approved'
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject="✅ Delegate Card Verified – Spandan 2025",
        message=f"Dear {reg.name},\nWe are delighted to confirm your registration as a delegate for Spandan 2025 - Beyond the Veil, scheduled to be held from August 25th to 30th at JIPMER, Puducherry. Your participation is a vital to making this event a success, and we are excited to welcome you to a vibrant lineup of activities and discussions!\n\n🪪 Registration Details:\n\n• Delegate Name: {reg.name}\n• College: {reg.college_name}\n• Tier: {reg.tier.upper()}\n• Delegate ID: {reg.user_id}\n• Date of Registration: {reg.created_at.strftime("%m/%d/%Y")}\n\nYou can use your **Delegate ID** `{reg.user_id}` to complete event registration through our official website.\n\nPlease carry a copy of this email and your college ID at the venue for smooth entry. Event guidelines and schedules will be shared soon.\n\nFor help, contact us at jsa.jipmer@gmail.com.\nWe look forward forward to hosting you at Spandan 2025!\n\nWarm regards,\nSuriya\nPresident, JIPMER Student Association"
    )
    return Response({"message": "Delegate card verified"}, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def reject_delegate_registration(request, pk):
    try:
        reg = DelegateCardRegistration.objects.get(pk=pk)
    except DelegateCardRegistration.DoesNotExist:
        return Response({'error': 'Registration not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.status = 'rejected'
    reg.is_verified = False
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject="❌ Delegate Card Rejected – Spandan 2025",
        message=f"Hi {reg.name},\n\nYour delegate card registration (Tier: {reg.tier}) has been rejected.\nPlease check your submission and try again.\n\nFor any query, contact us at jsa.jipmer@gmail.com.\n\nRegards,\nTeam Spandan"
    )
    return Response({'message': 'Delegate card rejected'}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def export_verified_delegate_cards(request):
    regs = DelegateCardRegistration.objects.filter(is_verified=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delegate Cards"
    headers = ["User ID", "Name", "Email", "Phone", "College", "Tier", "Amount", "Status", "Transaction ID", "Verified At", "Date"]
    ws.append(headers)
    for r in regs:
        ws.append([
            r.user_id, r.name, r.email, r.phone, r.college_name, r.tier,
            r.amount, r.status, r.transaction_id,
            r.verified_at.strftime("%Y-%m-%d %H:%M") if r.verified_at else "", 
            r.created_at.strftime("%Y-%m-%d")
        ])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=delegate_cards.xlsx'
    wb.save(response)
    return response

# Event Registration Views
class EventRegisterView(APIView):
    permission_classes = [permissions.AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request):
        regs = EventRegistration.objects.all()
        serializer = EventRegistrationSerializer(regs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def validate_delegate_ids(self, delegate_ids, email):
        if not delegate_ids:
            return True
        for delegate_id in delegate_ids:
            try:
                card = DelegateCardRegistration.objects.get(user_id=delegate_id, is_verified=True)
                if card.email != email:
                    return False
            except DelegateCardRegistration.DoesNotExist:
                return False
        return True

    def post(self, request):
        try:
            data = request.data.dict()  # Convert QueryDict to regular dict
            
            # Parse JSON fields
            json_fields = ['events', 'delegate_id']
            for field in json_fields:
                if field in data and isinstance(data[field], str):
                    try:
                        data[field] = json.loads(data[field])
                    except json.JSONDecodeError:
                        return Response(
                            {field: "Invalid JSON format"},
                            status=status.HTTP_400_BAD_REQUEST
                        )

            email = data.get("email", "").lower().strip()
            selected_events = data.get("events", [])
            
            if not isinstance(selected_events, list) or len(selected_events) == 0:
                return Response(
                    {"events": "At least one event must be selected"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validate delegate IDs if provided
            delegate_ids = data.get("delegate_id", [])
            if delegate_ids:
                if not isinstance(delegate_ids, list):
                    return Response(
                        {"delegate_id": "Must be an array"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                with transaction.atomic():
                    for delegate_id in delegate_ids:
                        if not DelegateCardRegistration.objects.filter(
                            user_id=delegate_id.strip(),
                            is_verified=True
                        ).exists():
                            return Response(
                                {"error": f"Delegate {delegate_id} not verified"},
                                status=status.HTTP_400_BAD_REQUEST
                            )

            # Check event requirements
            needs_delegate = any(
                event not in EXEMPT_EVENTS
                and not (
                    (event in SPORTS_EVENTS and PassPurchase.objects.filter(email=email, pass_type='sports', is_verified=True).exists()) or
                    (event in CULT_EVENTS and PassPurchase.objects.filter(email=email, pass_type='cult', is_verified=True).exists()) or
                    (event in LIT_EVENTS and PassPurchase.objects.filter(email=email, pass_type__in=['lit_lit', 'lit_premium'], is_verified=True).exists()) or
                    (event in QUIZ_EVENTS and PassPurchase.objects.filter(email=email, pass_type__in=['lit_quiz', 'lit_premium'], is_verified=True).exists())
                )
                for event in selected_events
            )

            if needs_delegate and not DelegateCardRegistration.objects.filter(email=email, is_verified=True).exists():
                return Response(
                    {"error": "Delegate card or valid pass required for selected events"},
                    status=status.HTTP_403_FORBIDDEN
                )
            EVENT_PRICE_MAP = {
                # Fine Arts
                'face_painting': 100,
                'pot_painting': 100,
                'mehendi': 100,
                'sketching': 100,
                'painting': 100,

                # Dance
                'solo_dance': 150,
                'duet_dance': 150,
                'chorea_theme': 250,
                'chorea_nontheme': 250,
                'show_down': 150,

                # Music
                'tinnitus': 250,
                'alaap': 250,
                'euphony': 150,
                'raag_rangmanch': 150,
                'solo_singing': 150,
                'solo_instrumental': 150,

                # Drama
                'play': 125,
                'skit': 125,
                'mime': 125,
                'adzap': 125,
                'variety': 125,
                'dernier_cri': 250,

                # Sports
                'cricket': 300,
                'football': 300,
                'basketball_men': 300,
                'basketball_women': 300,
                'volleyball_men': 300,
                'volleyball_women': 300,
                'hockey_men': 300,
                'hockey_women': 300,
                'futsal': 300,
                'chess_bullet': 200,
                'chess_rapid': 200,
                'chess_blitz': 200,
                'carroms': 150,
                'throwball_men': 300,
                'throwball_women': 300,
                'tennis': 300,
                'aquatics': 200,
                'badminton': 300,
                'table_tennis': 300,
                'athletics': 200,

                # Literary
                'malarkey': 150,
                'shipwrecked': 150,
                'turncoat': 150,
                'scrabble': 150,
                'formal_debate': 150,
                'cryptic_crossword': 150,
                'ppt_karaoke': 150,
                'potpourri': 150,
                'india_quiz': 150,
                'fandom_quiz': 150,
                'sports_quiz': 150,
                'rewind_quiz': 150,
                'formal_quiz': 250,
                'tj_jaishankar_memorial_quiz': 250,
                'jam': 250,
            }

            amount = sum(EVENT_PRICE_MAP.get(event, 0) for event in selected_events)
            data['amount'] = amount
            serializer = EventRegistrationSerializer(data=data)
            if serializer.is_valid():
                reg = serializer.save()
                
                # Auto-generate user_id if not set
                if not reg.user_id:
                    reg.user_id = f"USER-EV-{reg.id:05d}"
                    reg.save()
                
                return Response({
                    "message": "Event registration submitted",
                    "user_id": reg.user_id,
                    "registration": EventRegistrationSerializer(reg).data
                }, status=status.HTTP_201_CREATED)
            
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response(
                {"error": "Server error", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def verify_event_registration(request, pk):
    try:
        reg = EventRegistration.objects.get(pk=pk)
    except EventRegistration.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.is_verified = True
    reg.status = 'approved'
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject="✅ Event Registration Verified – Spandan 2025",
        message=f"Dear {reg.name},\n\nWe are thrilled to confirm your registration for the following events at Spandan 2025 - Beyond the Veil:\n{''.join(f"• {e}\n" for e in reg.events)}\n\n🧾 Registration Details:\nName: {reg.name}\nCollege: {reg.college}\nEmail: {reg.email}\nTotal Paid: ₹{reg.amount}\nEvent ID: {reg.user_id}\n• Date: {reg.created_at.strftime("%m/%d/%Y")}\n\nPlease carry a copy of this confirmation email and your delegate ID(if applicable) during the event.\n\nIf you have questions or need help, feel free to write to us at jsa.jipmer@gmail.com.\nAll the best and see you soon at Spandan 2025!\n\nWarm regards,\nTeam Spandan"
    )
    return Response({"message": "Event registration verified"}, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def reject_event_registration_soft(request, pk):
    try:
        reg = EventRegistration.objects.get(pk=pk)
    except EventRegistration.DoesNotExist:
        return Response({'error': 'Registration not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.status = 'rejected'
    reg.is_verified = False
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject="❌ Event Registration Rejected – Spandan 2025",
        message=f"Hi {reg.name},\n\nYour event registration has been rejected.\n\nEvents:\n" +
                "\n".join(f"• {e}" for e in reg.events) +
                "\n\nPlease review your payment and try again if needed.\n\nFor any query, contact us at jsa.jipmer@gmail.com.\n\nRegards,\nTeam Spandan"
    )
    return Response({'message': 'Event registration rejected'}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def export_verified_event_registrations(request):
    regs = EventRegistration.objects.filter(is_verified=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Event Registrations"
    headers = ["User ID", "Name", "Email", "Phone", "College", "Events", "Delegate IDs", "Amount", "Status", "Transaction ID", "Verified At", "Date"]
    ws.append(headers)
    for r in regs:
        ws.append([
            r.user_id, r.name, r.email, r.phone, r.college,
            ", ".join(r.events), ", ".join(r.delegate_id or []),
            r.amount, r.status, r.transaction_id,
            r.verified_at.strftime("%Y-%m-%d %H:%M") if r.verified_at else "", 
            r.created_at.strftime("%Y-%m-%d")
        ])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=event_registrations.xlsx'
    wb.save(response)
    return response

# Pass Purchase Views
class PassPurchaseView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = PassPurchaseSerializer(data=request.data)
        if serializer.is_valid():
            reg = serializer.save()
            return Response({
                "message": "Pass purchase submitted",
                "purchase": PassPurchaseSerializer(reg).data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        self.check_permissions(request)
        regs = PassPurchase.objects.all()
        serializer = PassPurchaseSerializer(regs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def verify_pass(request, pk):
    try:
        reg = PassPurchase.objects.get(pk=pk)
    except PassPurchase.DoesNotExist:
        return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.is_verified = True
    reg.status = 'approved'
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject=f"✅ {reg.get_pass_type_display()} Verified – Spandan 2025",
        message=f"Dear {reg.name},\n\nWe're excited to confirm that your {reg.pass_type} for Spandan2025 has been successfully verified!\n\n🎫 Pass Details:\n• Name: {reg.name}\n• College: {reg.college_name}\n• Pass: {reg.pass_type}\n• Pass ID: {reg.user_id}\n• Amount Paid: ₹{reg.amount}\n• Date of Purchase: {reg.created_at.strftime("%m/%d/%Y")}\n\nYour pass allows you to participate in eligible events under this category. Please carry this confirmation and your college ID for smooth verification at the venue.\n\nFor any support, feel free to reach us at jsa.jipmer@gmail.com.\n\nWarm regards,\nTeam Spandan"
    )
    return Response({"message": "Pass verified"}, status=status.HTTP_200_OK)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsAdminUser])
def reject_pass_soft(request, pk):
    try:
        reg = PassPurchase.objects.get(pk=pk)
    except PassPurchase.DoesNotExist:
        return Response({'error': 'Pass not found'}, status=status.HTTP_404_NOT_FOUND)

    reg.status = 'rejected'
    reg.is_verified = False
    reg.verified_at = timezone.now()
    reg.save()

    send_smtp_email(
        to_email=reg.email,
        subject=f"❌ {reg.get_pass_type_display()} Rejected – Spandan 2025",
        message=f"Hi {reg.name},\n\nYour pass purchase ({reg.get_pass_type_display()}) has been rejected.\nPlease check the screenshot or contact support.\n\nFor any query, contact us at jsa.jipmer@gmail.com.\n\nRegards,\nTeam Spandan"
    )
    return Response({'message': 'Pass rejected'}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def export_verified_passes(request):
    regs = PassPurchase.objects.filter(is_verified=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pass Purchases"
    headers = ["User ID", "Name", "Email", "Phone", "College", "Pass Type", "Amount", "Status", "Transaction ID", "Verified At", "Date"]
    ws.append(headers)
    for r in regs:
        ws.append([
            r.user_id, r.name, r.email, r.phone, r.college_name,
            r.get_pass_type_display(), r.amount, r.status, r.transaction_id,
            r.verified_at.strftime("%Y-%m-%d %H:%M") if r.verified_at else "", 
            r.created_at.strftime("%Y-%m-%d")
        ])
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pass_purchases.xlsx'
    wb.save(response)
    return response

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def check_delegate_eligibility(request, email):
    email = email.lower()
    try:
        reg = DelegateCardRegistration.objects.get(email=email)
        return Response({
            'is_verified': reg.is_verified,
            'user_id': reg.user_id,
            'tier': reg.tier
        })
    except DelegateCardRegistration.DoesNotExist:
        return Response({'is_verified': False})
    except DelegateCardRegistration.MultipleObjectsReturned:
        regs = DelegateCardRegistration.objects.filter(email=email)
        verified = any(reg.is_verified for reg in regs)
        return Response({
            'is_verified': verified,
            'user_id': [reg.user_id for reg in regs],
            'tier': [reg.tier for reg in regs]
        })